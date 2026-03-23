#!/usr/bin/env python3
"""
L&K Agency Transfer Market Analysis Generator
Generates Excel (.xlsx) and interactive HTML dashboard from scraped transfer market data.

Usage:
    python generate_files.py <fa_data.json> <ce_data.json> [output_dir]

Input:
    fa_data.json  - Free agents data (JSON array)
    ce_data.json  - Contract expiring players data (JSON array)
    output_dir    - Optional output directory (default: current directory)
"""

import json
import re
import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_market_value(value_str):
    """Convert market value string (e.g., 'â¬50.00m', 'â¬900k') to numeric value"""
    if not value_str or value_str == 'â¬0':
        return 0
    value_str = value_str.replace('â¬', '').strip()
    if 'm' in value_str.lower():
        return float(value_str.lower().replace('m', '')) * 1_000_000
    if 'k' in value_str.lower():
        return float(value_str.lower().replace('k', '')) * 1_000
    try:
        return float(value_str)
    except:
        return 0


def load_json_data(fa_file, ce_file):
    """Load player data from JSON files"""
    with open(fa_file, 'r', encoding='utf-8') as f:
        fa_players = json.load(f)
    with open(ce_file, 'r', encoding='utf-8') as f:
        ce_players = json.load(f)
    return fa_players, ce_players


def create_excel_file(fa_players, ce_players, output_path):
    """Create professionally formatted Excel file with two sheets"""
    wb = Workbook()
    wb.remove(wb.active)

    fa_sheet = wb.create_sheet("Free Agents")
    ce_sheet = wb.create_sheet("Contract Expiring 2026")

    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    alt_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- Free Agents Sheet ---
    fa_headers = ["Name", "Position", "Age", "Nationality", "Out of Contract Since", "Market Value", "Last League", "Profile URL"]
    for col, header in enumerate(fa_headers, 1):
        cell = fa_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    fa_sorted = sorted(fa_players, key=lambda x: parse_market_value(x.get('marketValue', 'â¬0')), reverse=True)
    for row_idx, player in enumerate(fa_sorted, 2):
        values = [
            player.get('name', ''), player.get('position', ''), player.get('age', ''),
            player.get('nationality', ''), player.get('outOfContractSince', ''),
            player.get('marketValue', 'â¬0'), player.get('lastLeague', ''), player.get('profileUrl', '')
        ]
        for col, value in enumerate(values, 1):
            cell = fa_sheet.cell(row=row_idx, column=col)
            cell.value = value
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if col == len(values) and value:
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # --- Contract Expiring Sheet ---
    ce_headers = ["Name", "Position", "Age", "Nationality", "Current Club", "League", "Market Value", "Profile URL"]
    for col, header in enumerate(ce_headers, 1):
        cell = ce_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ce_sorted = sorted(ce_players, key=lambda x: parse_market_value(x.get('marketValue', 'â¬0')), reverse=True)
    for row_idx, player in enumerate(ce_sorted, 2):
        values = [
            player.get('name', ''), player.get('position', ''), player.get('age', ''),
            player.get('nationality', ''), player.get('club', ''), player.get('league', ''),
            player.get('marketValue', 'â¬0'), player.get('profileUrl', '')
        ]
        for col, value in enumerate(values, 1):
            cell = ce_sheet.cell(row=row_idx, column=col)
            cell.value = value
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if col == len(values) and value:
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-fit columns + freeze header
    for sheet in [fa_sheet, ce_sheet]:
        for column in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
        sheet.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Excel file created: {output_path}")


def create_html_dashboard(fa_players, ce_players, output_path):
    """Create self-contained interactive HTML dashboard with dark theme"""

    for p in fa_players:
        p['numericValue'] = parse_market_value(p.get('marketValue', 'â¬0'))
    for p in ce_players:
        p['numericValue'] = parse_market_value(p.get('marketValue', 'â¬0'))

    fa_sorted = sorted(fa_players, key=lambda x: x['numericValue'], reverse=True)
    ce_sorted = sorted(ce_players, key=lambda x: x['numericValue'], reverse=True)

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>L&K Agency Transfer Dashboard 2026</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#0a0e17;color:#e0e0e0}}
.header{{background:linear-gradient(135deg,#0d1b2a,#1b2838);padding:24px 32px;border-bottom:2px solid #1e90ff}}
.header h1{{font-size:24px;color:#fff;font-weight:700}}
.header .sub{{color:#8899aa;font-size:13px;margin-top:4px}}
.stats{{display:flex;gap:16px;margin-top:12px}}
.stat{{background:rgba(30,144,255,0.1);border:1px solid rgba(30,144,255,0.3);border-radius:8px;padding:10px 18px}}
.stat .num{{font-size:22px;font-weight:700;color:#1e90ff}}
.stat .lbl{{font-size:11px;color:#8899aa;text-transform:uppercase}}
.tabs{{display:flex;gap:0;padding:0 32px;background:#0d1117}}
.tab{{padding:14px 28px;cursor:pointer;color:#8899aa;font-weight:600;border-bottom:3px solid transparent;transition:all 0.2s}}
.tab.active{{color:#1e90ff;border-bottom-color:#1e90ff}}
.tab:hover{{color:#fff}}
.controls{{padding:16px 32px;display:flex;gap:12px;flex-wrap:wrap;align-items:center}}
input[type="text"],select{{background:#161b22;border:1px solid #30363d;color:#e0e0e0;padding:8px 14px;border-radius:6px;font-size:13px}}
input[type="text"]{{width:280px}}
input[type="text"]:focus,select:focus{{outline:none;border-color:#1e90ff}}
.content{{padding:0 32px 32px}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
thead{{position:sticky;top:0;z-index:10}}
th{{background:#161b22;color:#8899aa;text-align:left;padding:10px 12px;font-weight:600;text-transform:uppercase;font-size:11px;cursor:pointer;border-bottom:2px solid #1e90ff;white-space:nowrap;user-select:none}}
th:hover{{color:#1e90ff}}
td{{padding:9px 12px;border-bottom:1px solid #1e2028}}
tr:nth-child(even){{background:rgba(255,255,255,0.02)}}
tr:hover{{background:rgba(30,144,255,0.08)}}
a{{color:#58a6ff;text-decoration:none}}
a:hover{{text-decoration:underline}}
.table-wrap{{max-height:calc(100vh - 260px);overflow-y:auto}}
.hidden{{display:none}}
.count{{color:#8899aa;font-size:12px;margin-left:auto}}
</style>
</head>
<body>
<div class="header">
<h1>L&K Agency Transfer Market Dashboard 2026</h1>
<div class="sub">Transfermarkt Data | Free Agents & Contract Expiring Players</div>
<div class="stats">
<div class="stat"><div class="num">{len(fa_sorted)}</div><div class="lbl">Free Agents</div></div>
<div class="stat"><div class="num">{len(ce_sorted)}</div><div class="lbl">Contract Expiring</div></div>
<div class="stat"><div class="num">{len(fa_sorted)+len(ce_sorted)}</div><div class="lbl">Total Players</div></div>
</div></div>
<div class="tabs">
<div class="tab active" data-tab="fa">Free Agents</div>
<div class="tab" data-tab="ce">Contract Expiring 2026</div>
</div>
<div class="controls">
<input type="text" id="search" placeholder="Search name, position, nationality...">
<select id="posFilter"><option value="">All Positions</option></select>
<select id="leagueFilter"><option value="">All Leagues</option></select>
<div class="count" id="rowCount"></div>
</div>
<div class="content">
<div id="fa-panel" class="panel"><div class="table-wrap">
<table><thead><tr><th data-col="0">Name</th><th data-col="1">Position</th><th data-col="2">Age</th><th data-col="3">Nationality</th><th data-col="4">Out of Contract</th><th data-col="5">Market Value</th><th data-col="6">League</th></tr></thead>
<tbody id="fa-body"></tbody></table></div></div>
<div id="ce-panel" class="panel hidden"><div class="table-wrap">
<table><thead><tr><th data-col="0">Name</th><th data-col="1">Position</th><th data-col="2">Age</th><th data-col="3">Nationality</th><th data-col="4">Club</th><th data-col="5">League</th><th data-col="6">Market Value</th></tr></thead>
<tbody id="ce-body"></tbody></table></div></div>
</div>
<script>
const faData={json.dumps(fa_sorted, ensure_ascii=False)};
const ceData={json.dumps(ce_sorted, ensure_ascii=False)};
let currentTab='fa';
function init(){{
  const tabs=document.querySelectorAll('.tab');
  tabs.forEach(t=>t.addEventListener('click',()=>{{
    tabs.forEach(x=>x.classList.remove('active'));t.classList.add('active');
    currentTab=t.dataset.tab;
    document.querySelectorAll('.panel').forEach(p=>p.classList.add('hidden'));
    document.getElementById(currentTab+'-panel').classList.remove('hidden');
    populateFilters();filterRows();
  }}));
  document.getElementById('search').addEventListener('input',filterRows);
  document.getElementById('posFilter').addEventListener('change',filterRows);
  document.getElementById('leagueFilter').addEventListener('change',filterRows);
  populateFilters();renderAll();filterRows();
}}
function populateFilters(){{
  const data=currentTab==='fa'?faData:ceData;
  const pos=new Set(),leagues=new Set();
  data.forEach(p=>{{if(p.position)pos.add(p.position);const l=currentTab==='fa'?p.lastLeague:p.league;if(l)leagues.add(l);}});
  const pf=document.getElementById('posFilter'),lf=document.getElementById('leagueFilter');
  pf.innerHTML='<option value="">All Positions</option>';
  lf.innerHTML='<option value="">All Leagues</option>';
  [...pos].sort().forEach(p=>{{const o=document.createElement('option');o.value=p;o.textContent=p;pf.appendChild(o);}});
  [...leagues].sort().forEach(l=>{{const o=document.createElement('option');o.value=l;o.textContent=l;lf.appendChild(o);}});
}}
function esc(s){{return(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')}}
function renderAll(){{
  const fab=document.getElementById('fa-body'),ceb=document.getElementById('ce-body');
  fab.innerHTML=faData.map(p=>'<tr><td><a href="https://www.transfermarkt.com'+esc(p.profileUrl)+'" target="_blank">'+esc(p.name)+'</a></td><td>'+esc(p.position)+'</td><td>'+p.age+'</td><td>'+esc(p.nationality)+'</td><td>'+esc(p.outOfContractSince)+'</td><td data-sort="'+p.numericValue+'">'+esc(p.marketValue)+'</td><td>'+esc(p.lastLeague)+'</td></tr>').join('');
  ceb.innerHTML=ceData.map(p=>'<tr><td><a href="https://www.transfermarkt.com'+esc(p.profileUrl)+'" target="_blank">'+esc(p.name)+'</a></td><td>'+esc(p.position)+'</td><td>'+p.age+'</td><td>'+esc(p.nationality)+'</td><td>'+esc(p.club)+'</td><td>'+esc(p.league)+'</td><td data-sort="'+p.numericValue+'">'+esc(p.marketValue)+'</td></tr>').join('');
}}
function filterRows(){{
  const q=document.getElementById('search').value.toLowerCase();
  const pos=document.getElementById('posFilter').value;
  const league=document.getElementById('leagueFilter').value;
  const body=document.getElementById(currentTab+'-body');
  let shown=0,total=0;
  body.querySelectorAll('tr').forEach(r=>{{
    total++;const text=r.textContent.toLowerCase();const cells=r.querySelectorAll('td');
    const pM=!pos||cells[1].textContent===pos;
    const lCol=currentTab==='fa'?6:5;const lM=!league||cells[lCol].textContent===league;
    const sM=!q||text.includes(q);
    if(pM&&lM&&sM){{r.style.display='';shown++;}}else{{r.style.display='none';}}
  }});
  document.getElementById('rowCount').textContent=shown+' / '+total+' players';
}}
document.querySelectorAll('th[data-col]').forEach(th=>{{
  th.addEventListener('click',()=>{{
    const table=th.closest('table'),tbody=table.querySelector('tbody');
    const rows=[...tbody.querySelectorAll('tr')],col=parseInt(th.dataset.col);
    const isAsc=th.classList.contains('sa');
    table.querySelectorAll('th').forEach(h=>{{h.classList.remove('sa','sd');}});
    th.classList.add(isAsc?'sd':'sa');
    rows.sort((a,b)=>{{
      let va=a.querySelectorAll('td')[col],vb=b.querySelectorAll('td')[col];
      if(va.dataset.sort!==undefined)return isAsc?parseFloat(vb.dataset.sort)-parseFloat(va.dataset.sort):parseFloat(va.dataset.sort)-parseFloat(vb.dataset.sort);
      va=va.textContent;vb=vb.textContent;
      const na=parseFloat(va),nb=parseFloat(vb);
      if(!isNaN(na)&&!isNaN(nb))return isAsc?nb-na:na-nb;
      return isAsc?vb.localeCompare(va):va.localeCompare(vb);
    }});
    rows.forEach(r=>tbody.appendChild(r));
  }});
}});
init();
</script>
</body></html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"HTML dashboard created: {output_path}")


def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_files.py <fa_data.json> <ce_data.json> [output_dir]")
        sys.exit(1)

    fa_file = sys.argv[1]
    ce_file = sys.argv[2]
    output_dir = sys.argv[3] if len(sys.argv) > 3 else "."

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    excel_output = os.path.join(output_dir, "LK_Agency_Transfer_Market_Analysis_2026.xlsx")
    html_output = os.path.join(output_dir, "LK_Agency_Transfer_Dashboard_2026.html")

    print("Loading data files...")
    fa_players, ce_players = load_json_data(fa_file, ce_file)
    print(f"Loaded {len(fa_players)} free agents and {len(ce_players)} contract expiring players")

    print("\nGenerating Excel file...")
    create_excel_file(fa_players, ce_players, excel_output)

    print("Generating HTML dashboard...")
    create_html_dashboard(fa_players, ce_players, html_output)

    print(f"\nFiles generated successfully!")
    print(f"  Excel: {excel_output}")
    print(f"  HTML:  {html_output}")


if __name__ == "__main__":
    main()
